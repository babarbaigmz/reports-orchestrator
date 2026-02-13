import csv
import itertools

from pathlib import Path
from typing import List, Optional, Generator, Union
from openpyxl import Workbook, load_workbook

from config.config import Config
from common.database_operations import QueryResults
from common.constants import FileModes, FileFormat, FileFormats
from common.exceptions import FileProcessingError


class FileWriter:
    """
    Handles writing data to CSV, Excel, or text files.
    Supports QueryResults, lists, generators, dicts, and tuples.
    """

    def __init__(self, config=Config):
        self.config = config
        self.__logger = self.config.get_logger()

    @staticmethod
    def ensure_path(filepath) -> Path:
        if isinstance(filepath, str):
            filepath = Path(filepath)
        return filepath.resolve()

    def __prepare_fieldnames_and_rows(self, file_path: Path, fieldnames, rows):
        """
        Public interface to write QueryResults, lists, or generators to CSV/Excel.
        """
        rows_iterator = iter(rows)

        if isinstance(rows, QueryResults):
            fieldnames = rows.fieldnames
            print("QueryResults")
        else:
            try:
                first_row = next(rows_iterator)
                rows_iterator = itertools.chain([first_row], rows_iterator)

                if fieldnames is None and isinstance(first_row, dict):
                    fieldnames = list(first_row.keys())
            except StopIteration:
                self.__logger.info(f"No data to write to the file: {file_path}")
                return

        return fieldnames, rows_iterator

    def write_excel(self,
                    file_path: Union[str, Path],
                    rows: Union[QueryResults, Generator[Union[dict, tuple], None, None], List[Union[dict, tuple]]],
                    write_by_cell: bool = True,
                    **kwargs
                    ):
        file_path = FileWriter.ensure_path(file_path)
        fieldnames = kwargs.get('fieldnames')
        fieldnames, rows_iterator = self.__prepare_fieldnames_and_rows(file_path, fieldnames, rows)
        kwargs.setdefault('fieldnames', fieldnames)

        if write_by_cell:
            self.__write_excel_by_cell(file_path, rows_iterator, **kwargs)
        else:
            self.__write_excel_bulk(file_path, rows_iterator, **kwargs)

    # --------------------------------------------------------- Excel Writing -------------------------------------------------------------------------#
    def __write_excel_by_cell(self,
                              file_path: Path,
                              rows_iterator: Generator[Union[dict, tuple], None, None],
                              **kwargs
                              ):

        self.__write_excel(file_path, rows_iterator, write_by_cell=True, **kwargs)

    def __write_excel_bulk(self,
                           file_path: Path,
                           rows_iterator: Generator[Union[dict, tuple], None, None],
                           **kwargs
                           ):

        self.__write_excel(file_path, rows_iterator, write_by_cell=False, **kwargs)

    # Helper to check if sheet is untouched
    def __is_sheet_empty(self, worksheet):
        return worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(1, 1).value is None

    def __write_excel(self,
                      file_path: Path,
                      rows: Generator[Union[dict, tuple], None, None],
                      write_by_cell: bool = True,
                      **kwargs
                      ):

        workbook = kwargs.get('workbook', None)
        fieldnames: Optional[List[str]] = kwargs.get('fieldnames')
        start_row: int = kwargs.get('start_row', 1)
        start_col: int = kwargs.get('start_col', 1)
        sheetname: str = kwargs.get('sheetname', 'Sheet1')
        mode: str = kwargs.get('mode', FileModes.MODE_WRITE.value)
        write_header: bool = True if mode == FileModes.MODE_WRITE.value else kwargs.get('write_header', True)
        batch_size: int = kwargs.get('batch_size', 1000)
        is_macro_file = file_path.suffix.lstrip(".") == FileFormat.FORMAT_EXCLM.value
        save: bool = kwargs.get('save', True)

        try:
            if workbook is None:
                if mode == FileModes.MODE_APPEND.value:
                    workbook = load_workbook(file_path, keep_vba=is_macro_file)
                else:
                    workbook = Workbook()
                    worksheet = workbook.active
                    worksheet.title = sheetname

            if sheetname in workbook.sheetnames:
                worksheet = workbook[sheetname]
            else:
                if len(workbook.sheetnames) == 1 and self.__is_sheet_empty(
                        workbook.active) and workbook.active.title.startswith("Sheet"):
                    worksheet = workbook.active
                    worksheet.title = sheetname
                else:
                    worksheet = workbook.create_sheet(sheetname)

            if write_header:
                for i, field in enumerate(fieldnames, start=start_col):
                    worksheet.cell(row=1, column=i, value=field)
                start_row += 1

            buffer = []
            rows_counter = 0

            for row in rows:
                row_values = [row.get(f) for f in fieldnames] if isinstance(row, dict) else row
                buffer.append(row_values)

                if len(buffer) == batch_size:
                    self.__flush_excel_buffer(worksheet, buffer, start_row, start_col, write_by_cell)
                    rows_counter += len(buffer)
                    self.__logger.info(f"{rows_counter} rows written to the file {file_path}")
                    start_row += len(buffer)
                    buffer.clear()

            if buffer:
                self.__flush_excel_buffer(worksheet, buffer, start_row, start_col, write_by_cell)
                rows_counter += len(buffer)
                self.__logger.info(f"{rows_counter} rows written to the file {file_path}")
                start_row += len(buffer)

            self.__logger.info(f"Export completed. Total: {rows_counter} rows written to the file {file_path}")

        except Exception as error:
            raise FileProcessingError(f"Failed to write data to {file_path} file") from error

        finally:
            if workbook and save:
                self.__logger.info(f"Saving file {file_path}")
                workbook.save(file_path)
                workbook.close()

    @staticmethod
    def __flush_excel_buffer(worksheet, buffer: list, start_row: int, start_col: int, write_by_cell: bool):
        if write_by_cell:
            for row in buffer:
                for i, val in enumerate(row, start=start_col):
                    worksheet.cell(row=start_row, column=i, value=val)
                start_row += 1
        else:
            for row in buffer:
                worksheet.append(row)

    def __write_query_results(self,
                              file_path: Union[str, Path],
                              rows: Union[
                                  'QueryResults', Generator[Union[dict, tuple], None, None], List[Union[dict, tuple]]],
                              write_by_cell: bool = True,
                              **kwargs
                              ):
        """
        Public interface to write QueryResults, lists, or generators to CSV/Excel.
        """
        file_path = FileWriter.ensure_path(file_path)

        if isinstance(rows, QueryResults):
            fieldnames = rows.fieldnames
            rows_iterator = iter(rows)
        else:
            rows_iterator = iter(rows)
            fieldnames = kwargs.get('fieldnames')

            if fieldnames is None:
                try:
                    first_row = next(rows_iterator)
                    rows_iterator = itertools.chain([first_row], rows_iterator)

                    if isinstance(first_row, dict):
                        fieldnames = list(first_row.keys())
                    else:
                        raise ValueError("Fieldnames must be provided")

                except StopIteration:
                    self.__logger.info(f"No data to write to {file_path}")
                    return

        kwargs.setdefault('fieldnames', fieldnames)

        file_extension = file_path.suffix.lstrip(".")

        if file_extension in FileFormats.text_extensions():
            self.__write_csv(file_path, rows_iterator, **kwargs)
        elif file_extension in FileFormats.excel_extensions():
            if write_by_cell:
                self.__write_excel_by_cell(file_path, rows_iterator, **kwargs)
            else:
                self.__write_excel_bulk(file_path, rows_iterator, **kwargs)
        else:
            raise NotImplementedError(f"Format not supported:")

    def write_csv_txt(self):
        pass

    def __write_csv(self,
                    file_path: Path,
                    rows: Union[Generator[dict, None, None], list],
                    **kwargs) -> None:
        """Write SQL statement results to a csv / txt file using DictWriter."""
        try:
            if isinstance(rows, list):
                if not rows:
                    self.__logger.info(f"No data to write to the file {file_path}")
                    return
            rows = iter(rows)

            try:
                first_row = next(rows)
                rows = itertools.chain([first_row], rows)
            except StopIteration:
                self.__logger.info(f"No data to write to the file {file_path}")
                return

            fieldnames: Optional[List[str]] = kwargs.get('fieldnames')
            delimiter: str = kwargs.get('delimiter', ',')
            mode: str = kwargs.get('mode', FileModes.MODE_WRITE.value)
            write_header = False if mode == FileModes.MODE_APPEND.value else kwargs.get('write_header', True)
            quote_option: str = kwargs.get('quote_option', 'm')
            batch_size: int = kwargs.get('batch_size', 10000)

            quote_mapping = {'m': csv.QUOTE_MINIMAL,
                             'a': csv.QUOTE_ALL,
                             'n': csv.QUOTE_NONE
                             }
            quoting = quote_mapping.get(quote_option, csv.QUOTE_MINIMAL)
            escape_char = '\\' if quote_option == 'n' else None

            file_data = []
            rows_counter = 0
            writer = None
            is_csv = file_path.suffix.lstrip(".") == FileFormat.FORMAT_CSV.value

            with open(file_path, mode=mode, newline='', encoding='utf-8') as file:

                for row in rows:
                    file_data.append(row)
                    if len(file_data) == batch_size:
                        writer, write_header, rows_counter = self.__write_data(file=file,
                                                                               file_data=file_data,
                                                                               fieldnames=fieldnames,
                                                                               writer=writer,
                                                                               write_header=write_header,
                                                                               rows_counter=rows_counter,
                                                                               is_csv=is_csv,
                                                                               quoting=quoting,
                                                                               delimiter=delimiter,
                                                                               escape_character=escape_char
                                                                               )

                        self.__logger.info(f"{rows_counter} rows written to the file {file_path}")

                if file_data:
                    writer, write_header, rows_counter = self.__write_data(
                        file=file,
                        file_data=file_data,
                        fieldnames=fieldnames,
                        writer=writer,
                        write_header=write_header,
                        rows_counter=rows_counter,
                        is_csv=is_csv,
                        quoting=quoting,
                        delimiter=delimiter,
                        escape_character=escape_char
                    )

                    self.__logger.info(f"Total {rows_counter} rows written to the file {file_path}")

                self.__logger.info(f"Export completed. {rows_counter} rows written to the file {file_path}")

        except Exception as error:
            raise FileProcessingError(f"Failed to write data to {file_path} file") from error
