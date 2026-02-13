import csv
import os
import shutil
import itertools

from pathlib import Path
from typing import List, Optional, Generator, Union, Tuple
from openpyxl import Workbook
from openpyxl import load_workbook

from config.config import Config
from common.constants import FileModes, FileFormat, FileFormats
from common.exceptions import FileProcessingError


class FileOperations:

    def __init__(self, config=Config):
        self.config = config
        self.__logger = self.config.get_logger()

    @staticmethod
    def ensure_path(filepath) -> Path:
        if isinstance(filepath, str):
            filepath = Path(filepath)
        return filepath.resolve()

    @property
    def logger(self):
        return self.__logger

    def __write_data(self,
                     file,
                     file_data: List[dict],
                     fieldnames: List[str],
                     writer: Optional[csv.writer],
                     write_header: bool,
                     rows_counter: int,
                     is_csv: bool,
                     quoting: int,
                     delimiter: str,
                     escape_character: str,
                     clear_file_data: bool = True) -> Tuple[Optional[csv.writer], bool, int]:
        if is_csv:
            if writer is None:
                writer = csv.DictWriter(file, fieldnames=fieldnames, quoting=quoting, delimiter=delimiter,
                                        escapechar=escape_character)

            if write_header:
                writer.writeheader()
                write_header = False

            writer.writerows(file_data)

        else:
            lines = []
            if write_header:
                lines.append(delimiter.join(fieldnames) + '\n')
                write_header = False

            for d in file_data:
                row_values = [str(d.get(f)) for f in fieldnames]
                lines.append(delimiter.join(row_values) + '\n')
            file.writelines(lines)

        rows_counter += len(file_data)

        if clear_file_data:
            file_data.clear()

        return writer, write_header, rows_counter

    def __write_csv(self,
                    file_path: Path,
                    rows: Union[Generator[dict, None, None], list],
                    **kwargs) -> None:
        """Write SQL statement results to a csv / txt file using DictWriter."""
        try:
            if isinstance(rows, list):
                if not rows:
                    self.__logger.info(f"No data to write to the file {file_path}")
                    return
            rows = iter(rows)

            try:
                first_row = next(rows)
                rows = itertools.chain([first_row], rows)
            except StopIteration:
                self.__logger.info(f"No data to write to the file {file_path}")
                return

            fieldnames: Optional[List[str]] = kwargs.get('fieldnames') or first_row.keys()

            delimiter: str = kwargs.get('delimiter', ',')
            mode: str = kwargs.get('mode', FileModes.MODE_WRITE.value)
            write_header = False if mode == FileModes.MODE_APPEND.value else kwargs.get('write_header', True)
            quote_option: str = kwargs.get('quote_option', 'm')
            batch_size: int = kwargs.get('batch_size', 10000)

            quote_mapping = {'m': csv.QUOTE_MINIMAL,
                             'a': csv.QUOTE_ALL,
                             'n': csv.QUOTE_NONE
                             }
            quoting = quote_mapping.get(quote_option, csv.QUOTE_MINIMAL)
            escape_char = '\\' if quote_option == 'n' else None

            file_data = []
            rows_counter = 0
            writer = None
            is_csv = file_path.suffix.lstrip(".") == FileFormat.FORMAT_CSV.value

            with open(file_path, mode=mode, newline='', encoding='utf-8') as file:

                for row in rows:
                    file_data.append(row)
                    if len(file_data) == batch_size:
                        writer, write_header, rows_counter = self.__write_data(file=file,
                                                                               file_data=file_data,
                                                                               fieldnames=fieldnames,
                                                                               writer=writer,
                                                                               write_header=write_header,
                                                                               rows_counter=rows_counter,
                                                                               is_csv=is_csv,
                                                                               quoting=quoting,
                                                                               delimiter=delimiter,
                                                                               escape_character=escape_char
                                                                               )

                        self.__logger.info(f"{rows_counter} rows written to the file {file_path}")

                if file_data:
                    writer, write_header, rows_counter = self.__write_data(file=file,
                                                                           file_data=file_data,
                                                                           fieldnames=fieldnames,
                                                                           writer=writer,
                                                                           write_header=write_header,
                                                                           rows_counter=rows_counter,
                                                                           is_csv=is_csv,
                                                                           quoting=quoting,
                                                                           delimiter=delimiter,
                                                                           escape_character=escape_char
                                                                           )

                    self.__logger.info(f"Total {rows_counter} rows written to the file {file_path}")

                self.__logger.info(f"Export completed. {rows_counter} rows written to the file {file_path}")

        except Exception as error:
            self.__logger.error(f"Failed to write data to {file_path} file due to error {error}", exc_info=True)
            raise

    def __write_excel(self,
                      file_path: Path,
                      rows: Union[Generator[dict, None, None], list],
                      **kwargs
                      ) -> None:
        """ Write SQL statement results to an Excel (.xlsx) file using openpyxl.Supports appending data starting at specified row and column. """

        workbook = None
        try:
            if isinstance(rows, list):
                if not rows:
                    self.__logger.info(f"No data to write to the file {file_path}")
                    return
            rows = iter(rows)
            try:
                first_row = next(rows)
                rows = itertools.chain([first_row], rows)

            except StopIteration:
                self.__logger.info(f"No data to write to the file {file_path}")
                return

            fieldnames: Optional[List[str]] = kwargs.get('fieldnames')
            mode: str = kwargs.get('mode', FileModes.MODE_APPEND.value)
            write_header: bool = False if mode == FileModes.MODE_APPEND.value else kwargs.get('write_header', False)

            if not fieldnames:
                if isinstance(first_row, dict):
                    fieldnames = list(first_row.keys())
                elif isinstance(first_row, (tuple, list)):
                    raise ValueError(
                        "Field names must be provided when rows are tuples/lists since column names cannot be inferred.")

            start_row: int = kwargs.get('start_row', 1)
            start_col: int = kwargs.get('start_col', 1)
            sheetname: str = kwargs.get('sheetname', 'Sheet1')
            batch_size: int = kwargs.get('batch_size', 10000)
            is_macro_file = file_path.suffix.lstrip(".") == FileFormat.FORMAT_EXCLM.value

            if mode == FileModes.MODE_APPEND.value:
                workbook = load_workbook(file_path, keep_vba=is_macro_file)
                if sheetname in workbook.sheetnames:
                    worksheet = workbook[sheetname]
                else:
                    worksheet = workbook.create_sheet(sheetname)
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = sheetname

            if write_header:
                for i, field in enumerate(fieldnames, start=start_col):
                    worksheet.cell(row=1, column=i, value=field)
                start_row += 1

            file_data = []
            rows_counter = 0

            for row in rows:
                file_data.append(row)
                if len(file_data) == batch_size:
                    for data in file_data:
                        for i, field in enumerate(fieldnames, start=start_col):
                            worksheet.cell(row=start_row, column=i, value=data.get(field))
                        start_row += 1
                    rows_counter += len(file_data)
                    self.__logger.info(f"{rows_counter} rows written to the file {file_path}")
                    file_data.clear()

            if file_data:
                for data in file_data:
                    for i, field in enumerate(fieldnames, start=start_col):
                        worksheet.cell(row=start_row, column=i, value=data.get(field))
                    start_row += 1
                rows_counter += len(file_data)
                self.__logger.info(f"{rows_counter} rows written to the file {file_path}")

            self.__logger.info(f"Export completed. {rows_counter} rows written to the file {file_path}")

        except Exception as error:
            raise FileProcessingError(f"Failed to write data to {file_path} file") from error

        finally:
            if workbook:
                workbook.save(file_path)

    def write_file(self, file_path: Path, rows: Generator[dict, None, None], **kwargs):
        file_path = FileOperations.ensure_path(file_path)
        file_extension = file_path.suffix.lstrip(".")

        if file_extension in FileFormats.text_extensions():
            self.__write_csv(file_path, rows, **kwargs)
        elif file_extension in FileFormats.excel_extensions():
            self.__write_excel(file_path, rows, **kwargs)
        else:
            raise NotImplementedError(f"Format not supported:")

    def read_file(self, file_path: Path, **kwargs):
        file_path = FileOperations.ensure_path(file_path)
        file_extension = file_path.suffix.lstrip(".")
        if file_extension in FileFormats.text_extensions():
            yield from self.__read_text(file_path, **kwargs)
        elif file_extension in FileFormats.excel_extensions():
            yield from self.__read_excel(file_path, **kwargs)
        else:
            raise NotImplementedError(f"Format not supported:")

    def __read_text(self, file_path: Path, **kwargs):
        try:
            file_path = FileOperations.ensure_path(file_path)
            delimiter: str = kwargs.get('delimiter', ',')
            skip_header: bool = kwargs.get('skip_header', False)
            self.__logger.info(f"Opening file: {file_path}")

            with open(file_path, "r", newline='', encoding="utf8") as file:
                reader = csv.reader(file, delimiter=delimiter)
                if skip_header:
                    next(reader, None)

                for row in reader:
                    stripped_row = [cell.strip() for cell in row]
                    if any(stripped_row):
                        yield stripped_row
                    else:
                        self.__logger.info("Skipping blank line.")

        except Exception as error:
            raise FileProcessingError(f"Failed to retrieve data from the file {file_path}") from error

    def __read_excel(self, file_path: Path, **kwargs):
        file_path = FileOperations.ensure_path(file_path)
        self.__logger.info(f"Opening file: {file_path}")

        start_row: int = kwargs.get('start_row', 1)
        start_col: int = kwargs.get('start_col', 1)
        max_col: int = kwargs.get('max_col', None)
        sheetname: str = kwargs.get('sheetname')

        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            if sheetname is None:
                sheetname = workbook.sheetnames[0]

            worksheet = workbook[sheetname] if isinstance(sheetname, str) else workbook.worksheets[sheetname]
            for row in worksheet.iter_rows(min_row=start_row, min_col=start_col, max_col=max_col, values_only=True):
                if any(row):
                    yield [cell for cell in row]

        except Exception as error:
            raise FileProcessingError(f"Failed to retrieve data from the file {file_path}") from error

        finally:
            if workbook:
                workbook.close()

    def copy_to_sharepoint(self, filepath: Path, sharepoint_folder: str, sharepoint_path: str):
        filepath = FileOperations.ensure_path(filepath)
        one_drive_location = os.environ.get("OneDrive")

        if not one_drive_location:
            raise FileProcessingError("OneDrive location unable to be found")

        sharepoint_mapped_location = Path(one_drive_location) / sharepoint_folder / filepath.name

        self.__logger.info(filepath)
        self.__logger.info(sharepoint_mapped_location)
        folder_path = sharepoint_folder.replace('\\', '/').replace(' ', '%20')
        sp_full_path = f"{sharepoint_path}/{folder_path}/{filepath}"
        shutil.copyfile(filepath, sharepoint_mapped_location)
        return sp_full_path

    def check_file(self, filepath: Path):
        return FileOperations.ensure_path(filepath).exists()

    def write_to_csv(self, filepath: Path, rows: List[dict], fieldnames: Optional[List[str]] = None,
                     mode: str = FileModes.MODE_WRITE.value, delimiter: str = ',', write_header: bool = True):
        """Write SQL statement results to a CSV file using DictWriter."""
        try:
            filepath = FileOperations.ensure_path(filepath)

            if not rows:
                self.__logger.warning(f"No data to write to the file {filepath}")
                return

            if fieldnames is None:
                fieldnames = list(rows[0].keys())

            with open(filepath, mode=mode, newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, quoting=csv.QUOTE_ALL, delimiter=delimiter)

                if write_header and mode == FileModes.MODE_WRITE.value:
                    writer.writeheader()

                for row in rows:
                    writer.writerow(row)

            self.__logger.info(f"Data written to CSV file: {filepath}")
        except Exception as error:
            raise FileProcessingError(f"Failed to write data to {filepath} file") from error

